from abc import ABC
import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path
import openpyxl

# not required but a good idea
class FileStrategy(ABC):
    def generate(self, fileName): pass

    def choose_file(self, name): pass

class CSVStrategy(FileStrategy):
    __file = None
    
    def __init__(self):
        self.__file = []
        
    def generate(self, fileName):
        pdf = FPDF(orientation="P", unit="mm", format="A4")

        for file in self.__file:
            df = pd.read_csv(file)

            for index, row in df.iterrows():
                pdf.add_page()
                
                pdf.set_font(family="Times", style="B", size=24)
                pdf.set_text_color(100, 100, 100)
                
                # add headers
                columns = df.columns
                columns = [item.replace("_", " ").title() for item in columns]
                
                pdf.set_font(family="Times", size=10, style="B")
                pdf.set_text_color(80, 80, 80)
                pdf.cell(w=30, h=8, txt=columns[0], border=1)
                pdf.cell(w=70, h=8, txt=columns[1], border=1)
                pdf.cell(w=30, h=8, txt=columns[2], border=1, ln=1)
                
                # add rows in table
                for index, row in df.iterrows():
                    pdf.set_font(family="Times", size=10)
                    pdf.set_text_color(80, 80, 80)
                    
                    #include str
                    pdf.cell(w=30, h=8, txt=str(row["student_id"]), border=1) 
                    pdf.cell(w=70, h=8, txt=str(row["student_name"]), border=1)
                    pdf.cell(w=30, h=8, txt=str(row["student_grade"]), border=1, ln=1)

        pdf.output(f"GeneratedPDFS/{fileName}.pdf")
    
    def choose_file(self, name):
        if (name.find(".csv") > -1):
            files = glob.glob(name)
            
            for file in files:
                self.__file.append(file)
                
        else:
            print("Not suited file type")

class TextStrategy(FileStrategy):

    __file = None

    def __init__(self):
        self.__file = []
        

    def generate(self, fileName):
        # Create one PDF file
        pdf = FPDF(orientation="P", unit="mm", format="A4")

        # Go through each text file
        for filepath in self.__file:
            # Add a page to the PDF document for each text file
            pdf.add_page()

            # Get the filename without the extension
            # and convert it to title case (e.g. Cat)
            filename = Path(filepath).stem
            name = filename.title()

            # Add the name to the PDF
            pdf.set_font(family="Times", size=16, style="B")
            pdf.cell(w=50, h=8, txt=name, ln=1)

            # Get the content of each text file
            with open(filepath, "r") as file:
                content = file.read()
            # Add the text file content to the PDf
            pdf.set_font(family="Times", size=12)
            pdf.multi_cell(w=0, h=6, txt=content)

        # Produce the PDF
        pdf.output(f"{fileName}.pdf")
     
    def choose_file(self, name):
        
        if (name.find(".txt") > -1):
            files = glob.glob(name)
            
            for file in files:
                self.__file.append(name)
        else:
            print("Not suited file type")


class ExcelStrategy(FileStrategy):
    
    __file = None
    
    def __init__(self):
        self.__file = []

    def generate(self, fileName):
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        for filepath in self.__file:    
            
            pdf.add_page()
            
            filename = Path(filepath).stem
            
            pdf.set_font(family="Times", size=16, style="B")
            pdf.cell(w=50, h=8, txt=f"Filename:{filename}", ln=1)
            
            # reading each excel file
            df = pd.read_excel(filepath)
            
            # add headers
            columns = df.columns
            columns = [item.replace("_", " ").title() for item in columns]
            
            pdf.set_font(family="Times", size=10, style="B")
            pdf.set_text_color(80, 80, 80)
            pdf.cell(w=30, h=8, txt=columns[0], border=1)
            pdf.cell(w=70, h=8, txt=columns[1], border=1)
            pdf.cell(w=30, h=8, txt=columns[2], border=1, ln=1)
            
            # add rows in table
            for index, row in df.iterrows():
                pdf.set_font(family="Times", size=10)
                pdf.set_text_color(80, 80, 80)
                
                #include str
                pdf.cell(w=30, h=8, txt=str(row["student_id"]), border=1) 
                pdf.cell(w=70, h=8, txt=str(row["student_name"]), border=1)
                pdf.cell(w=30, h=8, txt=str(row["student_grade"]), border=1, ln=1)
            
        pdf.output(f"GeneratedPDFS/{fileName}.pdf")

    def choose_file(self, name):
        if (name.find(".xlsx") > -1):
            files = glob.glob(name)
            
            for file in files:
                self.__file.append(file)
                
        else:
            print("Not suited file type")


if __name__ == '__main__':
    pass